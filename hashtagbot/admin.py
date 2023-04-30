import email
import imaplib
import random
import re

from django import forms
from django.contrib import admin
from django.http import HttpResponse
from django.template.context_processors import csrf
from django.urls import reverse
from django.utils.html import format_html
from instagrapi import Client
from instagrapi.mixins.challenge import ChallengeChoice
from openpyxl.workbook import Workbook

# Register your models here.

from hashtagbot.models import Worker, AccountInfo
def change_password_handler(username):
    # Simple way to generate a random string
    chars = list("abcdefghijklmnopqrstuvwxyz1234567890!&£@#")
    password = "".join(random.sample(chars, 8))
    return password
def get_code_from_sms(username):
    while True:
        code = input(f"Enter code (6 digits) for {username}: ").strip()
        if code and code.isdigit():
            return code
    return None


def challenge_code_handler(username, choice):
    if choice == ChallengeChoice.SMS:
        return get_code_from_sms(username)
    elif choice == ChallengeChoice.EMAIL:
        return get_code_from_email(username)
    return False

def get_code_from_email(username):
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(user="luccopilot@gmail.com",password= "knjkjbxusnapdtlo")
    mail.select("inbox")
    result, data = mail.search(None, "(UNSEEN)")
    assert result == "OK", "Error1 during get_code_from_email: %s" % result
    ids = data.pop().split()
    for num in reversed(ids):
        mail.store(num, "+FLAGS", "\\Seen")  # mark as read
        result, data = mail.fetch(num, "(RFC822)")
        assert result == "OK", "Error2 during get_code_from_email: %s" % result
        msg = email.message_from_string(data[0][1].decode())
        payloads = msg.get_payload()
        if not isinstance(payloads, list):
            payloads = [msg]
        code = None
        for payload in payloads:
            body = payload.get_payload(decode=True).decode()
            if "<div" not in body:
                continue
            match = re.search(">([^>]*?({u})[^<]*?)<".format(u=username), body)
            if not match:
                continue
            print("Match from email:", match.group(1))
            match = re.search(r">(\d{6})<", body)
            if not match:
                print('Skip this email, "code" not found')
                continue
            code = match.group(1)
            if code:
                return code
    return False

class WorkerAdmin(admin.ModelAdmin):
    list_display = ('username', 'is_active', 'is_working')
    list_filter = ('is_active', 'is_working')
    readonly_fields = ('user',)

    def save_model(self, request, obj, form, change):
        if not change:
            obj.user = request.user
            cl = Client()
            cl.change_password_handler = change_password_handler
            cl.challenge_code_handler = challenge_code_handler

            cl.login(obj.username, obj.password)


        obj.save()

    # filter query set by user
    def get_queryset(self, request):
        qs = super(WorkerAdmin, self).get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(user=request.user)


class AccountInfoAdmin(admin.ModelAdmin):
    change_list_template = "change_list.html"
    list_display = ('username', 'phone_number', 'from_tag', 'is_business', 'is_checked')
    list_filter = ('is_business', 'is_checked')
    readonly_fields = ('user',)

    def save_model(self, request, obj, form, change):
        if not change:
            obj.user = request.user
        obj.save()

    # filter query set by user
    def get_queryset(self, request):
        qs = super(AccountInfoAdmin, self).get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(user=request.user)

    def export_to_excel(self, request, queryset):
        # ساخت فایل اکسل
        wb = Workbook()
        ws = wb.active

        # افزودن ستون‌ها
        ws['A1'] = 'نام'
        ws['B1'] = 'شماره'

        # افزودن داده‌ها
        for obj in queryset:
            if obj.is_business:
                row = [obj.username, obj.phone_number]
                ws.append(row)


        # نام فایل اکسل
        filename = 'my_data.xlsx'

        # ذخیره فایل اکسل
        wb.save(filename)

        # ارسال فایل به کاربر
        with open(filename, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response







admin.site.register(Worker, WorkerAdmin)
admin.site.register(AccountInfo, AccountInfoAdmin)


admin.site.site_header = "نئون پدیا"
admin.site.site_title = "نئون پدیا"
admin.site.index_title = "پنل مدیریت ربات اینستاگرام"



